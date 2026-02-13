#!/usr/bin/env python3
"""
Debug: Verificar todas as abas do arquivo andradas
"""

import pandas as pd
from pathlib import Path
import openpyxl

filepath = Path('/home/agiliza/Documentos/gerador_planilhas/modelos_pedidos/Pedido labotrat  andradas.xlsx')

print(f"Analisando: {filepath.name}\n")

# Verificar abas disponíveis
print("="*80)
print("ABAS DISPONÍVEIS NO ARQUIVO:")
print("="*80)

wb = openpyxl.load_workbook(filepath)
print(f"Abas: {wb.sheetnames}\n")

# Ler cada aba
for sheet_name in wb.sheetnames:
    print(f"\n{'─'*80}")
    print(f"ABA: {sheet_name}")
    print(f"{'─'*80}")
    
    df = pd.read_excel(filepath, header=None, sheet_name=sheet_name)
    print(f"Dimensões: {len(df)} linhas x {len(df.columns)} colunas")
    
    if len(df) > 0:
        print("\nPrimeiras 10 linhas:")
        for idx in range(min(10, len(df))):
            valores = []
            for v in df.iloc[idx]:
                if pd.notna(v):
                    valores.append(str(v)[:50])
            if valores:  # Só mostra se tem valor
                print(f"  Linha {idx}: {' | '.join(valores)}")
    
    # Procurar CNPJ
    import re
    cnpj_pattern = r'\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}'
    
    found_cnpj = False
    for row_idx, row in df.iterrows():
        for col_idx, val in enumerate(row):
            if pd.notna(val):
                val_str = str(val).strip()
                if re.search(cnpj_pattern, val_str):
                    match = re.search(cnpj_pattern, val_str)
                    cnpj = match.group(0)
                    print(f"\n✅ CNPJ ENCONTRADO (Linha {row_idx+1}, Col {col_idx}): {cnpj}")
                    found_cnpj = True
    
    if not found_cnpj:
        print("\n❌ Nenhum CNPJ encontrado nesta aba")
